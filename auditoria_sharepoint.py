"""
=========================================================
 AUDITORIA DE SHAREPOINT - APP STREAMLIT (Microsoft Graph + Delta)
=========================================================
Instalacion:
    pip install streamlit msal requests pandas openpyxl

Uso:
    streamlit run app_auditoria_sharepoint.py

Esta version usa la API "delta" de Microsoft Graph: la primera vez
trae todo el contenido de cada carpeta configurada; en corridas
posteriores solo trae lo que cambio (nuevo / modificado / eliminado)
desde la ultima vez, usando un "deltaLink" guardado en disco.
"""

import streamlit as st
import msal
import requests
import pandas as pd
import json
import os
import re
from io import BytesIO
from urllib.parse import quote

# ===================== CONFIGURACION =====================

CLIENT_ID = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"  # Microsoft Azure CLI (first-party)
TENANT_ID = "aquanqape.onmicrosoft.com"
SP_HOSTNAME = "aquanqape.sharepoint.com"
GRAPH_URL = "https://graph.microsoft.com/v1.0"

CACHE_DIR = "cache_auditoria"

st.set_page_config(page_title="Auditoria SharePoint", page_icon="📊", layout="wide")


# ===================== UTILIDADES DE CACHE =====================

def _slug(texto):
    """Convierte texto arbitrario en un nombre de archivo seguro (solo alfanuméricos, guiones y underscores)."""
    return re.sub(r"[^a-zA-Z0-9_-]+", "_", texto).strip("_").lower()


def cache_paths(etiqueta):
    """Devuelve (ruta_csv, ruta_json_delta) para la etiqueta dada, creando el directorio si no existe."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    base = _slug(etiqueta)
    return (
        os.path.join(CACHE_DIR, f"{base}_data.csv"),
        os.path.join(CACHE_DIR, f"{base}_delta.json"),
    )


def cargar_cache(etiqueta):
    """Carga el DataFrame y el deltaLink guardados para la etiqueta; devuelve (None, None) si no hay cache."""
    data_path, delta_path = cache_paths(etiqueta)
    df = pd.read_csv(data_path) if os.path.exists(data_path) else None
    delta_link = None
    if os.path.exists(delta_path):
        with open(delta_path, "r", encoding="utf-8") as f:
            delta_link = json.load(f).get("delta_link")
    return df, delta_link


def guardar_cache(etiqueta, df, delta_link):
    """Persiste el DataFrame en CSV y el deltaLink en JSON para reutilizarlos en la próxima ejecución."""
    data_path, delta_path = cache_paths(etiqueta)
    df.to_csv(data_path, index=False)
    with open(delta_path, "w", encoding="utf-8") as f:
        json.dump({"delta_link": delta_link}, f)


def limpiar_cache(etiqueta):
    """Elimina los archivos de cache (CSV y JSON) asociados a la etiqueta para forzar una carga completa."""
    data_path, delta_path = cache_paths(etiqueta)
    for p in (data_path, delta_path):
        if os.path.exists(p):
            os.remove(p)


# ===================== FUNCIONES DE GRAPH =====================

def obtener_site_y_drive(headers, site_path):
    url = f"{GRAPH_URL}/sites/{SP_HOSTNAME}:{site_path}"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"No se pudo obtener el sitio '{site_path}': {resp.status_code} - {resp.text}")
    site_id = resp.json()["id"]

    url = f"{GRAPH_URL}/sites/{site_id}/drive"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"No se pudo obtener el drive: {resp.status_code} - {resp.text}")
    return site_id, resp.json()["id"]


def obtener_item_id(drive_id, item_path, headers):
    path_encoded = quote(item_path)
    url = f"{GRAPH_URL}/drives/{drive_id}/root:/{path_encoded}"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"No se encontro la carpeta '{item_path}': {resp.status_code} - {resp.text}")
    return resp.json()["id"]


def obtener_delta(drive_id, item_id, headers, delta_link=None):
    """Devuelve (items, nuevo_delta_link). Si delta_link es None, trae todo (full sync)."""
    url = delta_link or f"{GRAPH_URL}/drives/{drive_id}/items/{item_id}/delta"

    items = []
    data = {}
    while url:
        resp = requests.get(url, headers=headers)
        if resp.status_code != 200:
            raise Exception(f"Error en delta: {resp.status_code} - {resp.text}")
        data = resp.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")

    return items, data.get("@odata.deltaLink")


def item_a_fila(item, etiqueta, ruta_base):
    """Convierte un item de Graph en una fila de nuestro dataframe (o marca de eliminado)."""
    item_id = item["id"]

    if "deleted" in item:
        return {"ItemId": item_id, "_deleted": True}

    parent_path = item.get("parentReference", {}).get("path", "")
    parent_rel = parent_path.split(":", 1)[1] if ":" in parent_path else ""
    nombre = item.get("name", "")
    ruta = f"{parent_rel}/{nombre}".lstrip("/")

    # Ignorar el item raiz de la carpeta auditada (no aporta info util como fila)
    if ruta == ruta_base or nombre == "":
        return None

    created_by = item.get("createdBy", {}).get("user", {})
    modified_by = item.get("lastModifiedBy", {}).get("user", {})

    return {
        "ItemId": item_id,
        "_deleted": False,
        "Origen": etiqueta,
        "Ruta": ruta,
        "Nombre": nombre,
        "Tipo": "Carpeta" if "folder" in item else "Archivo",
        "Creado_Por": created_by.get("displayName", ""),
        "Email_Creador": created_by.get("email", ""),
        "Fecha_Creacion": item.get("createdDateTime", ""),
        "Modificado_Por": modified_by.get("displayName", ""),
        "Email_Modifico": modified_by.get("email", ""),
        "Fecha_Modificacion": item.get("lastModifiedDateTime", ""),
        "Tamano_Bytes": item.get("size", ""),
    }


def aplicar_cambios(df_actual, items, etiqueta, ruta_base):
    nuevas, eliminados = [], set()

    for item in items:
        fila = item_a_fila(item, etiqueta, ruta_base)
        if fila is None:
            continue
        if fila["_deleted"]:
            eliminados.add(fila["ItemId"])
        else:
            nuevas.append(fila)

    df_nuevas = pd.DataFrame(nuevas)
    if not df_nuevas.empty:
        df_nuevas = df_nuevas.drop(columns=["_deleted"])

    if df_actual is not None and not df_actual.empty:
        ids_actualizar = set(df_nuevas["ItemId"]) if not df_nuevas.empty else set()
        df_actual = df_actual[~df_actual["ItemId"].isin(eliminados)]
        df_actual = df_actual[~df_actual["ItemId"].isin(ids_actualizar)]
        df_final = pd.concat([df_actual, df_nuevas], ignore_index=True)
    else:
        df_final = df_nuevas

    return df_final, len(nuevas), len(eliminados)


def generar_excel(df):
    buffer = BytesIO()
    df_export = df.copy()

    for col in ["Fecha_Creacion", "Fecha_Modificacion"]:
        df_export[col] = pd.to_datetime(df_export[col], errors="coerce", utc=True).dt.tz_localize(None)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Auditoria")

        worksheet = writer.sheets["Auditoria"]
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

        for col_idx, col_name in enumerate(df_export.columns, start=1):
            max_len = max(
                df_export[col_name].astype(str).map(len).max() if len(df_export) else 0,
                len(col_name)
            )
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    buffer.seek(0)
    return buffer


# ===================== UI =====================

st.title("📊 Auditoría de SharePoint")
st.caption("Lista archivos y carpetas con quién los creó/modificó y cuándo. Las corridas siguientes solo traen lo que cambió.")

if "access_token" not in st.session_state:
    st.session_state.access_token = None

# ----- LOGIN -----
if st.session_state.access_token is None:
    st.subheader("1. Conectar con Microsoft 365")

    if st.button("🔑 Conectar"):
        app = msal.PublicClientApplication(
            client_id=CLIENT_ID,
            authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        )
        flow = app.initiate_device_flow(scopes=["https://graph.microsoft.com/.default"])

        if "user_code" not in flow:
            st.error(f"No se pudo iniciar el login: {flow}")
        else:
            st.info(f"👉 Abre **{flow['verification_uri']}** e ingresa el código:  \n\n## `{flow['user_code']}`")

            with st.spinner("Esperando que completes el login en el navegador..."):
                result = app.acquire_token_by_device_flow(flow)

            if "access_token" in result:
                st.session_state.access_token = result["access_token"]
                st.success("✅ Autenticado correctamente")
                st.rerun()
            else:
                st.error(f"Error de autenticación: {result.get('error_description')}")

else:
    st.success("✅ Conectado a Microsoft 365")
    if st.button("Cerrar sesión"):
        st.session_state.access_token = None
        st.rerun()

    st.divider()
    st.subheader("2. Configurar fuentes a auditar")

    if "fuentes" not in st.session_state:
        st.session_state.fuentes = pd.DataFrame([
            {
                "etiqueta": "Recepcion de Documentos - Oficinas Prize Peru",
                "site_path": "/sites/OficinasPrizePeru",
                "ruta_carpeta": "Recepción de Documentos",
            }
        ])

    fuentes_editadas = st.data_editor(
        st.session_state.fuentes,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "etiqueta": "Etiqueta / Origen",
            "site_path": "Ruta del sitio (ej: /sites/NombreSitio)",
            "ruta_carpeta": "Carpeta dentro de la biblioteca (ej: Recepción de Documentos)",
        },
    )

    st.divider()
    st.subheader("3. Generar / actualizar auditoría")

    col_a, col_b = st.columns(2)
    actualizar = col_a.button("🔄 Actualizar (solo cambios)", type="primary")
    completo = col_b.button("♻️ Reporte completo (reiniciar cache)")

    if actualizar or completo:
        headers = {
            "Authorization": f"Bearer {st.session_state.access_token}",
            "Accept": "application/json",
        }

        dfs_finales = []

        for _, fuente in fuentes_editadas.iterrows():
            etiqueta = fuente["etiqueta"]
            ruta_carpeta = fuente["ruta_carpeta"]

            if completo:
                limpiar_cache(etiqueta)

            df_cache, delta_link = cargar_cache(etiqueta)

            with st.spinner(f"[{etiqueta}] {'Trayendo todo (primera vez / reinicio)...' if delta_link is None else 'Consultando cambios...'}"):
                try:
                    site_id, drive_id = obtener_site_y_drive(headers, fuente["site_path"])
                    item_id = obtener_item_id(drive_id, ruta_carpeta, headers)
                    items, nuevo_delta_link = obtener_delta(drive_id, item_id, headers, delta_link)
                    df_actualizado, n_nuevos, n_eliminados = aplicar_cambios(df_cache, items, etiqueta, ruta_carpeta)
                    guardar_cache(etiqueta, df_actualizado, nuevo_delta_link)
                except Exception as e:
                    st.error(f"[{etiqueta}] Error: {e}")
                    continue

            if delta_link is None:
                st.info(f"[{etiqueta}] Primera carga: {len(df_actualizado)} items.")
            else:
                st.info(f"[{etiqueta}] Cambios detectados: {n_nuevos} nuevos/modificados, {n_eliminados} eliminados. Total actual: {len(df_actualizado)} items.")

            dfs_finales.append(df_actualizado)

        if dfs_finales:
            st.session_state.df_resultado = pd.concat(dfs_finales, ignore_index=True)

    # ----- RESULTADOS -----
    if "df_resultado" in st.session_state:
        df = st.session_state.df_resultado

        st.divider()
        st.subheader("4. Resultados")

        col1, col2, col3 = st.columns(3)
        col1.metric("Total items", len(df))
        col2.metric("Carpetas", (df["Tipo"] == "Carpeta").sum())
        col3.metric("Archivos", (df["Tipo"] == "Archivo").sum())

        usuarios = ["(todos)"] + sorted(df["Creado_Por"].dropna().unique().tolist())
        usuario_filtro = st.selectbox("Filtrar por creador", usuarios)

        df_mostrar = df if usuario_filtro == "(todos)" else df[df["Creado_Por"] == usuario_filtro]

        st.dataframe(df_mostrar, use_container_width=True, height=400)

        excel_buffer = generar_excel(df)
        st.download_button(
            "⬇️ Descargar Excel completo",
            data=excel_buffer,
            file_name="auditoria_sharepoint.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )